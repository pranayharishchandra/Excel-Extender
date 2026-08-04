"""
core/formula_engine.py

Core orchestration engine for the Excel Formula Extension Engine.

The FormulaEngine coordinates the complete workbook processing flow:

    Read Configuration
            ↓
        Validate
            ↓
    Detect Materials
            ↓
    Build Sheet Contexts
            ↓
    Execute Command
            ↓
      Save Workbook
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from commands.base_command import BaseCommand
from interfaces.logger import ILogger
from interfaces.progress import IProgress
from models.processing_result import ProcessingResult
from models.sheet_context import SheetContext
from services.config_reader import ConfigReader
from services.material_detector import MaterialDetector
from services.validator import Validator


class FormulaEngine:
    """
    Main orchestration engine for workbook processing.
    """

    def __init__(
        self,
        config_reader: ConfigReader,
        validator: Validator,
        material_detector: MaterialDetector,
        logger: ILogger,
        progress: IProgress,
    ) -> None:
        self._config_reader = config_reader
        self._validator = validator
        self._material_detector = material_detector
        self._logger = logger
        self._progress = progress

    def process(
        self,
        workbook_path: Path,
        command: BaseCommand,
        target_material_count: int,
        selected_sheets: list[str],
        output_path: Path,
    ) -> ProcessingResult:
        """
        Process the workbook.

        Args:
            workbook_path:
                Source workbook.

            command:
                Command to execute.

            target_material_count:
                Desired total material count.

            selected_sheets:
                Worksheets selected by the user.

            output_path:
                Destination workbook.

        Returns:
            ProcessingResult.
        """
        result = ProcessingResult()

        try:
            workbook = load_workbook(
                workbook_path,
                data_only=False,
                keep_vba=True,
            )

            configs = self._config_reader.read(workbook)

            validation = self._validator.validate_workbook(
                workbook,
                configs,
            )

            if not validation.is_valid:
                for error in validation.errors:
                    self._logger.error(error)

                result.set_error(
                    "Workbook validation failed."
                )
                return result

            selected_configs = [
                config
                for config in configs
                if config.sheet_name in selected_sheets
            ]

            self._progress.start(len(selected_configs))

            for step, config in enumerate(selected_configs, start=1):
                worksheet = workbook[config.sheet_name]

                previous_material_count = (
                    self._material_detector.detect_material_count(
                        worksheet,
                        config,
                    )
                )

                previous_last_row = (
                    self._material_detector.detect_last_material_row(
                        worksheet,
                        config,
                    )
                )

                target_last_row = (
                    config.template_start_row
                    + (
                        (target_material_count - 1)
                        * config.rows_per_material
                    )
                )

                context = SheetContext(
                    config=config,
                    previous_material_count=previous_material_count,
                    target_material_count=target_material_count,
                    previous_last_row=previous_last_row,
                    target_last_row=target_last_row,
                )

                self._progress.update(
                    step,
                    context.sheet_name,
                )

                command.execute(
                    worksheet,
                    context,
                )

                result.add_processed_sheet(
                    context.sheet_name,
                )

            workbook.save(output_path)

            self._progress.finish()

            self._logger.info(
                "Workbook saved successfully."
            )

            result.add_message(
                f"Saved workbook to '{output_path}'."
            )

            return result

        except Exception as exc:
            self._logger.error(str(exc))
            result.set_error(str(exc))
            return result

    @staticmethod
    def load_workbook(
        workbook_path: Path,
    ) -> Workbook:
        """
        Load an Excel workbook.

        Args:
            workbook_path:
                Workbook path.

        Returns:
            Loaded Workbook.
        """
        return load_workbook(
            workbook_path,
            data_only=False,
            keep_vba=True,
        )