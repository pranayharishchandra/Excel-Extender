"""
services/config_reader.py

Reads worksheet processing configuration from the
'pranay_extension_config' worksheet.

Configuration is identified by header names rather than fixed
column positions.
"""

from __future__ import annotations

from typing import List

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from common.column_utils import parse_managed_columns
from common.constants import (
    CONFIG_SHEET_NAME,
    FALSE_VALUES,
    TRUE_VALUES,
)
from common.excel_headers import (
    build_header_map,
    get_header_column,
    has_all_required_headers,
    missing_required_headers,
)
from models.sheet_config import SheetConfig


class ConfigReader:
    """
    Reads and validates worksheet configuration from the
    pranay_extension_config worksheet.
    """

    def read(self, workbook: Workbook) -> List[SheetConfig]:
        """
        Read all worksheet configurations.

        Args:
            workbook:
                Loaded workbook.

        Returns:
            List of SheetConfig objects.

        Raises:
            ValueError:
                If the configuration sheet or required headers
                are missing, or if configuration values are invalid.
        """
        worksheet = self._get_configuration_sheet(workbook)
        header_map = self._build_header_map(worksheet)

        configs: List[SheetConfig] = []

        for row in range(2, worksheet.max_row + 1):
            sheet_name = worksheet.cell(
                row=row,
                column=get_header_column(header_map, "Sheet Name"),
            ).value

            # Skip completely empty rows.
            if sheet_name is None or str(sheet_name).strip() == "":
                continue

            template_start_row = worksheet.cell(
                row=row,
                column=get_header_column(
                    header_map,
                    "Template Start Row",
                ),
            ).value

            rows_per_material = worksheet.cell(
                row=row,
                column=get_header_column(
                    header_map,
                    "Rows per Material",
                ),
            ).value

            managed_columns = worksheet.cell(
                row=row,
                column=get_header_column(
                    header_map,
                    "Managed Columns",
                ),
            ).value

            preserve_existing = worksheet.cell(
                row=row,
                column=get_header_column(
                    header_map,
                    "Preserve Existing",
                ),
            ).value

            configs.append(
                SheetConfig(
                    sheet_name=str(sheet_name).strip(),
                    template_start_row=self._parse_positive_int(
                        template_start_row,
                        "Template Start Row",
                        row,
                    ),
                    rows_per_material=self._parse_positive_int(
                        rows_per_material,
                        "Rows per Material",
                        row,
                    ),
                    managed_columns=parse_managed_columns(
                        "" if managed_columns is None else str(managed_columns)
                    ),
                    preserve_existing=self._parse_boolean(
                        preserve_existing
                    ),
                )
            )

        if not configs:
            raise ValueError(
                "No worksheet configurations were found in "
                f"'{CONFIG_SHEET_NAME}'."
            )

        return configs

    @staticmethod
    def _get_configuration_sheet(workbook: Workbook) -> Worksheet:
        """
        Retrieve the configuration worksheet.
        """
        if CONFIG_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Configuration sheet '{CONFIG_SHEET_NAME}' was not found."
            )

        return workbook[CONFIG_SHEET_NAME]

    @staticmethod
    def _build_header_map(worksheet: Worksheet) -> dict[str, int]:
        """
        Read the first row and build the header map.
        """
        headers = [
            cell.value if cell.value is not None else ""
            for cell in worksheet[1]
        ]

        header_map = build_header_map(headers)

        if not has_all_required_headers(header_map):
            missing = ", ".join(missing_required_headers(header_map))
            raise ValueError(
                f"Missing required configuration headers: {missing}"
            )

        return header_map

    @staticmethod
    def _parse_positive_int(
        value: object,
        field_name: str,
        row: int,
    ) -> int:
        """
        Parse a positive integer configuration value.
        """
        try:
            parsed = int(value)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Invalid '{field_name}' value on configuration row {row}."
            ) from exc

        if parsed < 1:
            raise ValueError(
                f"'{field_name}' must be greater than zero "
                f"(configuration row {row})."
            )

        return parsed

    @staticmethod
    def _parse_boolean(value: object) -> bool:
        """
        Parse the Preserve Existing value.

        Accepted values are defined in common.constants.
        """
        if value is None:
            return False

        normalized = str(value).strip().upper()

        if normalized in TRUE_VALUES:
            return True

        if normalized in FALSE_VALUES:
            return False

        raise ValueError(
            "Invalid value for 'Preserve Existing'. "
            "Expected one of: "
            f"{', '.join(sorted(TRUE_VALUES | FALSE_VALUES))}"
        )