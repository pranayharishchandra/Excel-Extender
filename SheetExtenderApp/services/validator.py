"""
services/validator.py

Validation service for the Excel Formula Extension Engine.

This service validates the workbook structure and configuration before any
processing begins.
"""

from __future__ import annotations

from openpyxl.workbook.workbook import Workbook

from common.constants import CONFIG_SHEET_NAME
from models.sheet_config import SheetConfig
from models.validation_result import ValidationResult


class Validator:
    """
    Performs workbook and configuration validation.
    """

    def validate_workbook(
        self,
        workbook: Workbook,
        sheet_configs: list[SheetConfig],
    ) -> ValidationResult:
        """
        Validate the workbook before processing.

        Validation performed:

        - Configuration sheet exists.
        - At least one SheetConfig exists.
        - Every configured worksheet exists.
        - Configuration sheet is not duplicated.
        - Managed columns are present.
        - Template row is inside worksheet bounds.

        Args:
            workbook:
                Loaded workbook.

            sheet_configs:
                Parsed worksheet configurations.

        Returns:
            ValidationResult.
        """
        result = ValidationResult()

        self._validate_configuration_sheet(
            workbook,
            result,
        )

        self._validate_sheet_configs(
            sheet_configs,
            result,
        )

        self._validate_configured_sheets(
            workbook,
            sheet_configs,
            result,
        )

        return result

    @staticmethod
    def _validate_configuration_sheet(
        workbook: Workbook,
        result: ValidationResult,
    ) -> None:
        """
        Validate that the configuration worksheet exists.
        """
        if CONFIG_SHEET_NAME not in workbook.sheetnames:
            result.add_error(
                f"Configuration worksheet '{CONFIG_SHEET_NAME}' was not found."
            )

    @staticmethod
    def _validate_sheet_configs(
        sheet_configs: list[SheetConfig],
        result: ValidationResult,
    ) -> None:
        """
        Validate the parsed configuration objects.
        """
        if not sheet_configs:
            result.add_error("No worksheet configuration was found.")
            return

        seen: set[str] = set()

        for config in sheet_configs:
            if config.sheet_name in seen:
                result.add_error(
                    f"Duplicate configuration for worksheet "
                    f"'{config.sheet_name}'."
                )
            else:
                seen.add(config.sheet_name)

            if not config.managed_columns:
                result.add_error(
                    f"Worksheet '{config.sheet_name}' "
                    f"does not contain any managed columns."
                )

            if config.template_start_row < 1:
                result.add_error(
                    f"Worksheet '{config.sheet_name}' has an invalid "
                    f"Template Start Row."
                )

            if config.rows_per_material < 1:
                result.add_error(
                    f"Worksheet '{config.sheet_name}' has an invalid "
                    f"Rows per Material value."
                )

    @staticmethod
    def _validate_configured_sheets(
        workbook: Workbook,
        sheet_configs: list[SheetConfig],
        result: ValidationResult,
    ) -> None:
        """
        Validate each configured worksheet.
        """
        for config in sheet_configs:
            if config.sheet_name not in workbook.sheetnames:
                result.add_error(
                    f"Configured worksheet '{config.sheet_name}' "
                    f"does not exist."
                )
                continue

            worksheet = workbook[config.sheet_name]

            if worksheet.max_row < config.template_start_row:
                result.add_error(
                    f"Worksheet '{config.sheet_name}' "
                    f"does not contain Template Start Row "
                    f"{config.template_start_row}."
                )

            if worksheet.max_column == 0:
                result.add_warning(
                    f"Worksheet '{config.sheet_name}' is empty."
                )

            for column in config.managed_columns:
                cell = worksheet[f"{column}{config.template_start_row}"]

                if cell.value is None:
                    result.add_warning(
                        f"Template cell '{cell.coordinate}' "
                        f"in worksheet '{config.sheet_name}' is empty."
                    )