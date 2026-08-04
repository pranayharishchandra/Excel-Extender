"""
gui/app.py

Main GUI application for the Excel Formula Extension Engine.
"""

from __future__ import annotations

from pathlib import Path

import customtkinter as ctk
from openpyxl import load_workbook

from commands.clear_command import ClearCommand
from commands.clear_extend_command import ClearExtendCommand
from commands.extend_command import ExtendCommand
from common.constants import (
    APPLICATION_NAME,
    APPLICATION_VERSION,
    DEFAULT_WINDOW_HEIGHT,
    DEFAULT_WINDOW_WIDTH,
    MIN_WINDOW_HEIGHT,
    MIN_WINDOW_WIDTH,
)
from core.formula_engine import FormulaEngine
from gui.dialogs import Dialogs
from gui.logger import GuiLogger
from gui.widgets import (
    LabeledComboBox,
    LabeledEntry,
    ProgressPanel,
    SheetList,
)
from interfaces.progress import IProgress
from services.config_reader import ConfigReader
from services.formula_clearer import FormulaClearer
from services.formula_extender import FormulaExtender
from services.material_detector import MaterialDetector
from services.validator import Validator


class App(ctk.CTk, IProgress):
    """
    Main application window.
    """

    def __init__(self) -> None:
        super().__init__()

        self.title(f"{APPLICATION_NAME} {APPLICATION_VERSION}")
        self.geometry(f"{DEFAULT_WINDOW_WIDTH}x{DEFAULT_WINDOW_HEIGHT}")
        self.minsize(MIN_WINDOW_WIDTH, MIN_WINDOW_HEIGHT)

        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("blue")

        self._workbook_path: Path | None = None
        self._output_path: Path | None = None
        self._configs = []
        self._total_progress_steps = 0

        self._config_reader = ConfigReader()
        self._validator = Validator()
        self._material_detector = MaterialDetector()

        self._formula_clearer = FormulaClearer()
        self._formula_extender = FormulaExtender()

        self._engine: FormulaEngine | None = None
        self._logger: GuiLogger | None = None

        self._build_layout()
        self._create_engine()

    def _build_layout(self) -> None:
        """
        Create all UI controls.
        """
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)
        self.grid_rowconfigure(5, weight=2)

        self._build_workbook_section()
        self._build_options_section()
        self._build_sheet_section()
        self._build_action_section()
        self._build_progress_section()
        self._build_log_section()

    def _build_workbook_section(self) -> None:
        frame = ctk.CTkFrame(self)
        frame.grid(
            row=0,
            column=0,
            padx=10,
            pady=(10, 5),
            sticky="ew",
        )

        frame.grid_columnconfigure(0, weight=1)

        self.workbook_entry = LabeledEntry(
            frame,
            "Workbook",
        )
        self.workbook_entry.grid(
            row=0,
            column=0,
            padx=10,
            pady=10,
            sticky="ew",
        )
        self.workbook_entry.configure_state("disabled")

        browse_button = ctk.CTkButton(
            frame,
            text="Browse...",
            width=120,
            command=self._browse_workbook,
        )
        browse_button.grid(
            row=0,
            column=1,
            padx=(0, 10),
            pady=10,
        )

    def _build_options_section(self) -> None:
        frame = ctk.CTkFrame(self)
        frame.grid(
            row=1,
            column=0,
            padx=10,
            pady=5,
            sticky="ew",
        )

        frame.grid_columnconfigure((0, 1, 2), weight=1)

        self.previous_material_entry = LabeledEntry(
            frame,
            "Previous Material Count",
        )
        self.previous_material_entry.grid(
            row=0,
            column=0,
            padx=10,
            pady=10,
            sticky="ew",
        )
        self.previous_material_entry.configure_state("disabled")

        self.new_material_entry = LabeledEntry(
            frame,
            "New Material Count",
        )
        self.new_material_entry.grid(
            row=0,
            column=1,
            padx=10,
            pady=10,
            sticky="ew",
        )

        self.output_mode = LabeledComboBox(
            frame,
            "Output",
            values=[
                "Save Original",
                "Save As",
            ],
        )
        self.output_mode.grid(
            row=0,
            column=2,
            padx=10,
            pady=10,
            sticky="ew",
        )

    def _build_sheet_section(self) -> None:
        frame = ctk.CTkFrame(self)
        frame.grid(
            row=2,
            column=0,
            padx=10,
            pady=5,
            sticky="nsew",
        )

        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(1, weight=1)

        header = ctk.CTkFrame(frame, fg_color="transparent")
        header.grid(
            row=0,
            column=0,
            padx=10,
            pady=(10, 5),
            sticky="ew",
        )

        header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            header,
            text="Sheets",
            font=ctk.CTkFont(size=15, weight="bold"),
        ).grid(
            row=0,
            column=0,
            sticky="w",
        )

        select_all_button = ctk.CTkButton(
            header,
            text="Select All",
            width=100,
            command=self._select_all_sheets,
        )
        select_all_button.grid(
            row=0,
            column=1,
            padx=(5, 0),
        )

        select_none_button = ctk.CTkButton(
            header,
            text="Select None",
            width=100,
            command=self._select_no_sheets,
        )
        select_none_button.grid(
            row=0,
            column=2,
            padx=(5, 0),
        )

        self.sheet_list = SheetList(
            frame,
            height=180,
        )
        self.sheet_list.grid(
            row=1,
            column=0,
            padx=10,
            pady=(0, 10),
            sticky="nsew",
        )

    def _build_action_section(self) -> None:
        frame = ctk.CTkFrame(self)
        frame.grid(
            row=3,
            column=0,
            padx=10,
            pady=5,
            sticky="ew",
        )

        clear_button = ctk.CTkButton(
            frame,
            text="Clear",
            command=self._clear,
        )
        clear_button.pack(
            side="left",
            padx=10,
            pady=10,
        )

        extend_button = ctk.CTkButton(
            frame,
            text="Extend",
            command=self._extend,
        )
        extend_button.pack(
            side="left",
            padx=10,
            pady=10,
        )

        clear_extend_button = ctk.CTkButton(
            frame,
            text="Clear + Extend",
            command=self._clear_extend,
        )
        clear_extend_button.pack(
            side="left",
            padx=10,
            pady=10,
        )

    """=============================================================="""
    """                            part 2                            """
    """=============================================================="""

    def _build_progress_section(self) -> None:
        """
        Create the progress panel.
        """
        self.progress_panel = ProgressPanel(self)
        self.progress_panel.grid(
            row=4,
            column=0,
            padx=10,
            pady=5,
            sticky="ew",
        )

    def _build_log_section(self) -> None:
        """
        Create the scrolling log console.
        """
        frame = ctk.CTkFrame(self)
        frame.grid(
            row=5,
            column=0,
            padx=10,
            pady=(5, 10),
            sticky="nsew",
        )

        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(0, weight=1)

        self.log_textbox = ctk.CTkTextbox(
            frame,
            wrap="word",
        )
        self.log_textbox.grid(
            row=0,
            column=0,
            padx=10,
            pady=10,
            sticky="nsew",
        )

        self._logger = GuiLogger(self.log_textbox)

    def _create_engine(self) -> None:
        """
        Create the FormulaEngine instance.
        """
        self._engine = FormulaEngine(
            config_reader=self._config_reader,
            validator=self._validator,
            material_detector=self._material_detector,
            logger=self._logger,
            progress=self,
        )

    def _browse_workbook(self) -> None:
        """
        Handle workbook selection.
        """
        workbook_path = Dialogs.select_workbook()

        if workbook_path is None:
            return

        self._workbook_path = workbook_path

        self.workbook_entry.configure_state("normal")
        self.workbook_entry.set(str(workbook_path))
        self.workbook_entry.configure_state("disabled")

        self._load_workbook_information()

    def _load_workbook_information(self) -> None:
        """
        Load workbook metadata and populate the UI.
        """
        assert self._workbook_path is not None

        self._logger.clear()
        self.progress_panel.reset()

        self._logger.info("Loading workbook...")

        workbook = load_workbook(
            self._workbook_path,
            data_only=False,
            keep_vba=True,
        )

        self._configs = self._config_reader.read(workbook)

        validation = self._validator.validate_workbook(
            workbook,
            self._configs,
        )

        if not validation.is_valid:
            for error in validation.errors:
                self._logger.error(error)

            Dialogs.show_error(
                "Validation Failed",
                "\n".join(validation.errors),
            )
            return

        for warning in validation.warnings:
            self._logger.warning(warning)

        sheet_names: list[str] = []
        previous_counts: list[int] = []

        for config in self._configs:
            worksheet = workbook[config.sheet_name]

            count = self._material_detector.detect_material_count(
                worksheet,
                config,
            )

            previous_counts.append(count)
            sheet_names.append(config.sheet_name)

        self.sheet_list.set_sheets(sheet_names)

        detected_count = max(previous_counts) if previous_counts else 1

        self.previous_material_entry.configure_state("normal")
        self.previous_material_entry.set(str(detected_count))
        self.previous_material_entry.configure_state("disabled")

        self.new_material_entry.clear()
        self.new_material_entry.set(str(detected_count))

        self._logger.info(
            f"Detected {len(sheet_names)} configured worksheet(s)."
        )

        self._logger.info(
            f"Detected {detected_count} material(s)."
        )

    def _select_all_sheets(self) -> None:
        """
        Select every worksheet.
        """
        self.sheet_list.select_all()

    def _select_no_sheets(self) -> None:
        """
        Deselect every worksheet.
        """
        self.sheet_list.deselect_all()

    def _clear(self) -> None:
        """
        Execute the Clear command.
        """
        command = ClearCommand(
            formula_clearer=self._formula_clearer,
            logger=self._logger,
            progress=self,
        )

        self._execute(command)

    def _extend(self) -> None:
        """
        Execute the Extend command.
        """
        command = ExtendCommand(
            formula_extender=self._formula_extender,
            logger=self._logger,
            progress=self,
        )

        self._execute(command)

    def _clear_extend(self) -> None:
        """
        Execute the Clear + Extend command.
        """
        command = ClearExtendCommand(
            formula_clearer=self._formula_clearer,
            formula_extender=self._formula_extender,
            logger=self._logger,
            progress=self,
        )

        self._execute(command)


    """=============================================================="""
    """                            part 3                            """
    """=============================================================="""
    def _execute(
        self,
        command,
    ) -> None:
        """
        Execute the selected command.

        Args:
            command:
                Instance of a BaseCommand implementation.
        """
        if self._workbook_path is None:
            Dialogs.show_error(
                "Workbook Required",
                "Please select an Excel workbook.",
            )
            return

        selected_sheets = self.sheet_list.get_selected()

        if not selected_sheets:
            Dialogs.show_error(
                "No Worksheets Selected",
                "Please select at least one worksheet.",
            )
            return

        try:
            target_material_count = int(
                self.new_material_entry.get()
            )
        except ValueError:
            Dialogs.show_error(
                "Invalid Material Count",
                "New Material Count must be a positive integer.",
            )
            return

        if target_material_count < 1:
            Dialogs.show_error(
                "Invalid Material Count",
                "New Material Count must be greater than zero.",
            )
            return

        output_mode = self.output_mode.get()

        if output_mode == "Save Original":
            overwrite = Dialogs.confirm_save_original()

            if not overwrite:
                return

            output_path = self._workbook_path

        else:
            output_path = Dialogs.select_output_workbook(
                self._workbook_path,
            )

            if output_path is None:
                return

        self._output_path = output_path

        self._logger.clear()
        self.progress_panel.reset()

        self._logger.info("Processing workbook...")

        assert self._engine is not None

        result = self._engine.process(
            workbook_path=self._workbook_path,
            command=command,
            target_material_count=target_material_count,
            selected_sheets=selected_sheets,
            output_path=output_path,
        )

        if result.success:
            for message in result.messages:
                self._logger.info(message)

            Dialogs.show_info(
                "Completed",
                "Workbook processed successfully.",
            )
        else:
            if result.error:
                self._logger.error(result.error)

            Dialogs.show_error(
                "Processing Failed",
                result.error or "Unknown error.",
            )

    # ------------------------------------------------------------------
    # IProgress implementation
    # ------------------------------------------------------------------

    def start(
        self,
        total_steps: int,
    ) -> None:
        """
        Begin progress reporting.
        """
        self._total_progress_steps = max(1, total_steps)
        self.progress_panel.reset()

    def update(
        self,
        current_step: int,
        current_sheet: str,
    ) -> None:
        """
        Update progress during processing.

        Args:
            current_step:
                Current completed step.

            current_sheet:
                Worksheet currently being processed.
        """
        self.progress_panel.update_progress(
            current=current_step,
            total=self._total_progress_steps,
            sheet_name=current_sheet,
        )

        self.update_idletasks()


    """=============================================================="""
    """                            part 5                            """
    """=============================================================="""
    def finish(self) -> None:
        """
        Complete progress reporting.
        """
        self.progress_panel.update_progress(
            current=self._total_progress_steps,
            total=self._total_progress_steps,
            sheet_name="Completed",
        )

        self.update_idletasks()

    # ------------------------------------------------------------------
    # Utility Methods
    # ------------------------------------------------------------------

    def reset_ui(self) -> None:
        """
        Reset the UI to its initial state while preserving the
        currently selected workbook.
        """
        self.sheet_list.clear()
        self.progress_panel.reset()

        self.previous_material_entry.configure_state("normal")
        self.previous_material_entry.clear()
        self.previous_material_entry.configure_state("disabled")

        self.new_material_entry.clear()

        self._configs.clear()
        self._total_progress_steps = 0

        if self._logger is not None:
            self._logger.clear()

    def clear_workbook(self) -> None:
        """
        Clear the currently loaded workbook from the application.
        """
        self.reset_ui()

        self._workbook_path = None
        self._output_path = None

        self.workbook_entry.configure_state("normal")
        self.workbook_entry.clear()
        self.workbook_entry.configure_state("disabled")

    @property
    def workbook_path(self) -> Path | None:
        """
        Returns the currently selected workbook path.
        """
        return self._workbook_path

    @property
    def output_path(self) -> Path | None:
        """
        Returns the output workbook path from the last execution.
        """
        return self._output_path

    @property
    def selected_sheets(self) -> list[str]:
        """
        Returns the currently selected worksheet names.
        """
        return self.sheet_list.get_selected()

    @property
    def target_material_count(self) -> int:
        """
        Returns the target material count entered by the user.

        Returns:
            The requested material count.

        Raises:
            ValueError:
                If the value cannot be converted to an integer.
        """
        return int(self.new_material_entry.get())

    def run(self) -> None:
        """
        Start the application.
        """
        self.mainloop()


if __name__ == "__main__":
    app = App()
    app.run()
