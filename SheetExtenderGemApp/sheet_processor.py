"""Executes safe Excel mutations based on loaded configs and openpyxl logic."""

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from models import SheetConfig
from workbook_utils import calculate_last_row
from typing import Callable

class SheetProcessor:
    def __init__(self, sheet: Worksheet, config: SheetConfig, logger: Callable[[str], None]):
        self.sheet = sheet
        self.config = config
        self.logger = logger

    def clear(self, old_materials: int):
        """
        Clears ONLY managed columns from beneath the sacred template row down to the calculated old target.
        """
        if old_materials == 0:
            return
            
        old_last_row = calculate_last_row(
            self.config.template_start_row, old_materials, self.config.rows_per_material
        )
        # Template row is sacred. Never clear it.
        start_clear_row = self.config.template_start_row + 1

        if start_clear_row > old_last_row:
            return

        self.logger(f"Clearing {self.config.sheet_name} from row {start_clear_row} to {old_last_row}...")
        
        for row in range(start_clear_row, old_last_row + 1):
            for col_idx in self.config.managed_columns:
                self.sheet.cell(row=row, column=col_idx).value = None

    def extend(self, new_materials: int):
            """
            Generates formulas mapping references via openpyxl Translator, leaving formatting untouched.
            Copies constants directly. Applies strictly to Managed Columns.
            """
            if new_materials == 0:
                return
                
            new_last_row = calculate_last_row(
                self.config.template_start_row, new_materials, self.config.rows_per_material
            )
            start_extend_row = self.config.template_start_row + 1

            if start_extend_row > new_last_row:
                return

            self.logger(f"Extending {self.config.sheet_name} from row {start_extend_row} to {new_last_row}...")

            for col_idx in self.config.managed_columns:
                template_cell = self.sheet.cell(row=self.config.template_start_row, column=col_idx)
                template_value = template_cell.value
                
                if template_value is None:
                    continue

                origin_coord = f"{get_column_letter(col_idx)}{self.config.template_start_row}"
                is_formula = isinstance(template_value, str) and template_value.strip().startswith('=')

                for row in range(start_extend_row, new_last_row + 1):
                    target_cell = self.sheet.cell(row=row, column=col_idx)
                    
                    if is_formula:
                        target_coord = f"{get_column_letter(col_idx)}{row}"
                        
                        try:
                            translated = Translator(template_value, origin=origin_coord).translate_formula(target_coord)
                            
                            # Fix XML corruption: Assign formula string cleanly
                            target_cell.value = str(translated)
                            
                            # Clear stale cached formula evaluation values
                            if hasattr(target_cell, '_value'):
                                target_cell.data_type = 'f'
                                
                        except Exception as e:
                            self.logger(f"WARN: Translation failed at {target_coord}. Fallback to original formula. Error: {e}")
                            target_cell.value = str(template_value)
                            target_cell.data_type = 'f'
                    else:
                        target_cell.value = template_value




