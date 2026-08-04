"""Presentation layer utilizing CustomTkinter."""

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog
import os
import threading
import datetime
import openpyxl

import constants
from config_reader import read_configuration
from workbook_utils import detect_material_count
from formula_engine import FormulaExtensionEngine

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Excel Formula Extension Engine")
        self.geometry("700x850")
        self.minsize(650, 800)

        ctk.set_appearance_mode("Dark")
        ctk.set_default_color_theme("blue")

        self.file_path = ""
        self.sheet_vars = {}
        self.old_materials = 0

        self._build_ui()

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(6, weight=1)

        # --- 1. Workbook Layout ---
        wb_frame = ctk.CTkFrame(self)
        wb_frame.grid(row=0, column=0, padx=20, pady=10, sticky="ew")
        wb_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(wb_frame, text="Workbook", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        
        self.path_entry = ctk.CTkEntry(wb_frame, placeholder_text="Select SAP template workbook...")
        self.path_entry.grid(row=1, column=0, columnspan=2, padx=10, pady=5, sticky="ew")
        self.path_entry.configure(state="disabled")

        browse_btn = ctk.CTkButton(wb_frame, text="Browse", width=100, command=self._browse_workbook)
        browse_btn.grid(row=1, column=2, padx=10, pady=5)

        ctk.CTkLabel(wb_frame, text="Previous Materials:").grid(row=2, column=0, padx=10, pady=10, sticky="w")
        self.lbl_prev_mat = ctk.CTkLabel(wb_frame, text="0", font=ctk.CTkFont(weight="bold"))
        self.lbl_prev_mat.grid(row=2, column=1, padx=10, pady=10, sticky="w")

        ctk.CTkLabel(wb_frame, text="New Materials:").grid(row=2, column=1, padx=10, pady=10, sticky="e")
        self.entry_new_mat = ctk.CTkEntry(wb_frame, width=80)
        self.entry_new_mat.grid(row=2, column=2, padx=10, pady=10, sticky="e")

        # --- 2. Sheets Configuration ---
        sheets_frame = ctk.CTkFrame(self)
        sheets_frame.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        sheets_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(sheets_frame, text="Sheets", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")

        self.scroll_sheets = ctk.CTkScrollableFrame(sheets_frame, height=120)
        self.scroll_sheets.grid(row=1, column=0, columnspan=2, padx=10, pady=5, sticky="ew")

        btn_frame = ctk.CTkFrame(sheets_frame, fg_color="transparent")
        btn_frame.grid(row=2, column=0, columnspan=2, padx=10, pady=5, sticky="w")
        
        ctk.CTkButton(btn_frame, text="Select All", width=100, command=lambda: self._toggle_all(True)).pack(side="left", padx=(0, 10))
        ctk.CTkButton(btn_frame, text="Select None", width=100, command=lambda: self._toggle_all(False)).pack(side="left")

        # --- 3. Actions Target ---
        actions_frame = ctk.CTkFrame(self)
        actions_frame.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        actions_frame.grid_columnconfigure((0, 1, 2), weight=1)

        ctk.CTkLabel(actions_frame, text="Actions", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")

        self.btn_clear = ctk.CTkButton(actions_frame, text="Clear", command=lambda: self._run_action("Clear"))
        self.btn_clear.grid(row=1, column=0, padx=10, pady=10, sticky="ew")

        self.btn_extend = ctk.CTkButton(actions_frame, text="Extend", command=lambda: self._run_action("Extend"))
        self.btn_extend.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

        self.btn_clear_extend = ctk.CTkButton(actions_frame, text="Clear + Extend", command=lambda: self._run_action("Clear + Extend"))
        self.btn_clear_extend.grid(row=1, column=2, padx=10, pady=10, sticky="ew")

        # --- 4. Output Route ---
        output_frame = ctk.CTkFrame(self)
        output_frame.grid(row=3, column=0, padx=20, pady=10, sticky="ew")
        output_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(output_frame, text="Output", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")

        self.output_mode = tk.StringVar(value="Original")
        self.radio_orig = ctk.CTkRadioButton(output_frame, text="Save to Original Workbook", variable=self.output_mode, value="Original", command=self._toggle_output)
        self.radio_orig.grid(row=1, column=0, padx=10, pady=5, sticky="w", columnspan=2)

        self.radio_save_as = ctk.CTkRadioButton(output_frame, text="Save As", variable=self.output_mode, value="SaveAs", command=self._toggle_output)
        self.radio_save_as.grid(row=2, column=0, padx=10, pady=5, sticky="w")

        self.output_path_entry = ctk.CTkEntry(output_frame)
        self.output_path_entry.grid(row=3, column=0, columnspan=2, padx=10, pady=5, sticky="ew")
        self.output_path_entry.configure(state="disabled")

        self.btn_out_browse = ctk.CTkButton(output_frame, text="Browse", width=100, state="disabled", command=self._browse_output)
        self.btn_out_browse.grid(row=3, column=2, padx=10, pady=5)

        # --- 5. Progress ---
        progress_frame = ctk.CTkFrame(self)
        progress_frame.grid(row=4, column=0, padx=20, pady=10, sticky="ew")
        progress_frame.grid_columnconfigure(0, weight=1)

        self.progress_bar = ctk.CTkProgressBar(progress_frame)
        self.progress_bar.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        self.progress_bar.set(0)

        self.lbl_progress = ctk.CTkLabel(progress_frame, text="Idle")
        self.lbl_progress.grid(row=1, column=0, padx=10, pady=(0,5))

        # --- 6. Event Logs ---
        logs_frame = ctk.CTkFrame(self)
        logs_frame.grid(row=5, column=0, padx=20, pady=10, sticky="nsew")
        logs_frame.grid_columnconfigure(0, weight=1)
        logs_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(logs_frame, text="Logs", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        
        self.txt_logs = ctk.CTkTextbox(logs_frame, state="disabled", wrap="word")
        self.txt_logs.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

    def _log_safe(self, msg: str):
        """Thread-safe UI logging operation"""
        timestamp = datetime.datetime.now().strftime("%H:%M")
        log_line = f"{timestamp} {msg}\n"
        self.txt_logs.configure(state="normal")
        self.txt_logs.insert("end", log_line)
        self.txt_logs.see("end")
        self.txt_logs.configure(state="disabled")

    def _toggle_output(self):
        if self.output_mode.get() == "SaveAs":
            self.output_path_entry.configure(state="normal")
            self.btn_out_browse.configure(state="normal")
        else:
            self.output_path_entry.delete(0, "end")
            self.output_path_entry.configure(state="disabled")
            self.btn_out_browse.configure(state="disabled")

    def _browse_workbook(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not path:
            return
        
        self.file_path = path
        self.path_entry.configure(state="normal")
        self.path_entry.delete(0, "end")
        self.path_entry.insert(0, self.file_path)
        self.path_entry.configure(state="disabled")

        self._log_safe("Inspecting Workbook Metadata...")
        threading.Thread(target=self._load_workbook_metadata, daemon=True).start()

    def _load_workbook_metadata(self):
        """Background parser utilizing read_only to quickly fetch config constraints."""
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            
            if constants.CONFIG_SHEET_NAME not in wb.sheetnames:
                raise ValueError(f"Missing '{constants.CONFIG_SHEET_NAME}' config sheet")
                
            if constants.SOURCE_SHEET_NAME not in wb.sheetnames:
                raise ValueError(f"Missing '{constants.SOURCE_SHEET_NAME}' source sheet")

            configs = read_configuration(wb[constants.CONFIG_SHEET_NAME])
            old_mat = detect_material_count(wb[constants.SOURCE_SHEET_NAME])
            sheet_names = [c.sheet_name for c in configs]
            wb.close()

            # Schedule UI update thread-safely
            self.after(0, self._update_ui_after_load, sheet_names, old_mat)
        except Exception as e:
            self.after(0, self._log_safe, f"Error inspecting workbook: {str(e)}")

    def _update_ui_after_load(self, sheet_names, old_mat):
        self.old_materials = old_mat
        self.lbl_prev_mat.configure(text=str(old_mat))
        self._log_safe(f"Found {old_mat} previous materials.")

        for widget in self.scroll_sheets.winfo_children():
            widget.destroy()

        self.sheet_vars.clear()

        # Dynamically spawn config-driven checkboxes
        for sn in sheet_names:
            var = tk.BooleanVar(value=True)
            chk = ctk.CTkCheckBox(self.scroll_sheets, text=sn, variable=var)
            chk.pack(anchor="w", pady=2, padx=5)
            self.sheet_vars[sn] = var
            
        self._log_safe("Sheets populated from configuration.")

    def _toggle_all(self, state: bool):
        for var in self.sheet_vars.values():
            var.set(state)

    def _browse_output(self):
        default_dir = os.path.dirname(self.file_path) if self.file_path else ""
        path = filedialog.asksaveasfilename(
            initialdir=default_dir,
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if path:
            self.output_path_entry.configure(state="normal")
            self.output_path_entry.delete(0, "end")
            self.output_path_entry.insert(0, path)

    def _set_ui_state(self, state: str):
        self.btn_clear.configure(state=state)
        self.btn_extend.configure(state=state)
        self.btn_clear_extend.configure(state=state)
        self.btn_out_browse.configure(state=state if self.output_mode.get() == "SaveAs" else "disabled")

    def _run_action(self, action: str):
        if not self.file_path:
            self._log_safe("Error: No workbook selected.")
            return

        try:
            new_mat = int(self.entry_new_mat.get())
            if new_mat <= 0:
                raise ValueError
        except ValueError:
            self._log_safe("Error: Enter a valid positive number for New Materials.")
            return

        selected = [sn for sn, var in self.sheet_vars.items() if var.get()]
        if not selected:
            self._log_safe("Error: No sheets selected.")
            return

        if self.output_mode.get() == "SaveAs":
            out_path = self.output_path_entry.get()
            if not out_path:
                self._log_safe("Error: Select output path for Save As.")
                return
        else:
            out_path = self.file_path

        self._set_ui_state("disabled")
        self.progress_bar.set(0)
        self.lbl_progress.configure(text="Starting...")

        engine = FormulaExtensionEngine(
            file_path=self.file_path,
            output_path=out_path,
            new_materials=new_mat,
            selected_sheets=selected,
            action=action,
            log_callback=lambda m: self.after(0, self._log_safe, m),
            progress_callback=lambda p, s: self.after(0, self._update_progress, p, s),
            completion_callback=lambda: self.after(0, self._process_complete)
        )
        engine.run()

    def _update_progress(self, val: float, sheet: str):
        self.progress_bar.set(val)
        self.lbl_progress.configure(text=f"Current: {sheet}")

    def _process_complete(self):
        self._set_ui_state("normal")
        self.lbl_progress.configure(text="Idle")