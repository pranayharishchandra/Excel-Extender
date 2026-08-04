"""The core orchestration engine. Independent of the GUI framework."""

import threading
import openpyxl
import os
from typing import Callable, List
import constants
from config_reader import read_configuration
from workbook_utils import detect_material_count
from sheet_processor import SheetProcessor

class FormulaExtensionEngine:
    def __init__(self,
                 file_path: str,
                 output_path: str,
                 new_materials: int,
                 selected_sheets: List[str],
                 action: str,
                 log_callback: Callable[[str], None],
                 progress_callback: Callable[[float, str], None],
                 completion_callback: Callable[[], None]):
                 
        self.file_path = file_path
        self.output_path = output_path
        self.new_materials = new_materials
        self.selected_sheets = selected_sheets
        self.action = action
        self.log_callback = log_callback
        self.progress_callback = progress_callback
        self.completion_callback = completion_callback
        self.thread = None

    def run(self):
        """Dispatches the processing engine to a daemonized background thread."""
        self.thread = threading.Thread(target=self._process, daemon=True)
        self.thread.start()

    def _process(self):
        try:
            self.log_callback(f"Loading Workbook: {os.path.basename(self.file_path)}")
            self.progress_callback(0.0, "Loading Workbook")
            
            # data_only=False ensures we read formulas instead of cached values
            wb = openpyxl.load_workbook(self.file_path, data_only=False)
            self.log_callback("Workbook Loaded")

            if constants.CONFIG_SHEET_NAME not in wb.sheetnames:
                raise ValueError(f"Configuration sheet '{constants.CONFIG_SHEET_NAME}' not found.")
            if constants.SOURCE_SHEET_NAME not in wb.sheetnames:
                raise ValueError(f"Source sheet '{constants.SOURCE_SHEET_NAME}' not found.")

            config_sheet = wb[constants.CONFIG_SHEET_NAME]
            configs = read_configuration(config_sheet)
            self.log_callback("Config Loaded")

            source_sheet = wb[constants.SOURCE_SHEET_NAME]
            old_materials = detect_material_count(source_sheet)
            self.log_callback(f"Previous Materials Detected: {old_materials}")

            total_sheets = len(self.selected_sheets)
            for i, sheet_name in enumerate(self.selected_sheets):
                self.progress_callback(i / total_sheets, sheet_name)

                if sheet_name not in wb.sheetnames:
                    self.log_callback(f"Warning: Sheet {sheet_name} not found in workbook. Skipping.")
                    continue

                sheet_config = next((c for c in configs if c.sheet_name == sheet_name), None)
                if not sheet_config:
                    self.log_callback(f"Warning: No config found for {sheet_name}. Skipping.")
                    continue

                processor = SheetProcessor(wb[sheet_name], sheet_config, self.log_callback)

                # Route actions
                if self.action in ("Clear", "Clear + Extend"):
                    processor.clear(old_materials)

                if self.action in ("Extend", "Clear + Extend"):
                    processor.extend(self.new_materials)

                self.log_callback(f"{sheet_name} Completed")

            self.progress_callback(0.9, "Saving...")
            self.log_callback("Saving Workbook...")

            # Force Excel to rebuild the calculation chain on next load
            if wb.calculation:
                wb.calculation.calcMode = 'auto'
                
            wb.save(self.output_path)
            
            self.progress_callback(1.0, "Done")
            self.log_callback("Finished Successfully")

        except Exception as e:
            self.log_callback(f"ERROR: {str(e)}")
            self.progress_callback(0.0, "Error")
        finally:
            self.completion_callback()