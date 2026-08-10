from openpyxl import load_workbook, Workbook

file_path = r"C:\Users\PranayHarishchandra\Desktop\Excel Extender\FormulaExtracter\Pranay Master carton all plants.xlsx"

wb = load_workbook(file_path, data_only=False)

for ws in wb.worksheets:
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Formulas"

    for row in ws.iter_rows():
        for cell in row:
            new_ws[cell.coordinate] = str(cell.value)



    new_wb.save(f"{ws.title}_Formulas.xlsx")
    print(f"Created {ws.title}_Formulas.xlsx")












'''all sheets in one file'''

# from openpyxl import load_workbook, Workbook

# file_path = r"C:\Users\PranayHarishchandra\Desktop\Master DSM 10 lines.xlsx"

# # Load workbook (keep formulas)
# wb = load_workbook(file_path, data_only=False)

# # Create new workbook
# new_wb = Workbook()

# # Remove the default sheet
# new_wb.remove(new_wb.active)

# # Loop through all sheets
# for ws in wb.worksheets:
#     print(f"Processing: {ws.title}")

#     # Create a sheet with the same name
#     new_ws = new_wb.create_sheet(title=ws.title)

#     # Copy all cell values/formulas as text
#     for row in ws.iter_rows():
#         for cell in row:
#             new_ws[cell.coordinate] = str(cell.value)

# # Save the output workbook
# output_file = "AllSheets_Formulas.xlsx"
# new_wb.save(output_file)

# print(f"\nCreated {output_file}")